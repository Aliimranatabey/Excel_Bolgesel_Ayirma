import openpyxl

kaynak_dosya_adi = 'denemeMesh.xlsx'
hedef_dosya_adi = 'Mesh İhtiyacı Olabilecek Müşterilerin Dağılımları.xlsx'

kaynak_workbook = openpyxl.load_workbook(kaynak_dosya_adi)
hedef_workbook = openpyxl.load_workbook(hedef_dosya_adi)

kaynak_sheet = kaynak_workbook.active
veri_satirlari = list(kaynak_sheet.iter_rows(values_only=True))

# Bölgeye göre çalışma sayfalarını saklamak için bir sözlük oluşturun
bolge_sayfalari = {}
ilkSatir=["Bölge","İl","İlçe","Alt Yapı","Wi-Fi Versiyon","Profil","MAC","SERVICENO"]
for satir in veri_satirlari[1:]:
    bolge_degerleri = satir[0]
    
    if bolge_degerleri not in bolge_sayfalari:
        bolge_sayfalari[bolge_degerleri] = hedef_workbook.create_sheet(title=bolge_degerleri)
        bolge_sayfalari[bolge_degerleri].append(ilkSatir)

    
    bolge_sayfalari[bolge_degerleri].append(list(satir))

hedef_workbook.save(hedef_dosya_adi)
