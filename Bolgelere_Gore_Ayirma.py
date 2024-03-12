import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

kaynak_dosya_adi = 'Mesh İhtiyaci Olabilecek Musterilerin Dagılımlari Dolu.xlsx'
hedef_dosya_adi = 'Mesh İhtiyacı Olabilecek Müşterilerin Dağılımlari.xlsx'

kaynak_workbook = openpyxl.load_workbook(kaynak_dosya_adi)
hedef_workbook = openpyxl.load_workbook(hedef_dosya_adi)

kaynak_sheet = kaynak_workbook.active
veri_satirlari = list(kaynak_sheet.iter_rows(values_only=True))

# Bölgeye göre çalışma sayfalarını saklamak için bir sözlük oluşturun
bolge_sayfalari = {}

ilkSatir=["Bölge","İl","İlçe","Alt Yapı","Wi-Fi Versiyon","Profil","MAC","SERVICENO"]
# if kaynak_dosya_adi=='Hattı Sağlıklı Olup Saturasyon Yaşayan Upsell Adayı Müsteriler (DSL) Dolu.xlsx':
#             ilkSatir.remove("Profil")
for i in veri_satirlari:
    if "Bölge"==veri_satirlari[0][0]:
        veri_satirlari.remove(i)

for satir_index, satir in enumerate(veri_satirlari):
    bolge_degerleri = satir[0]
    
    if bolge_degerleri not in bolge_sayfalari:
        bolge_sayfalari[bolge_degerleri] = hedef_workbook.create_sheet(title=bolge_degerleri)
        
        bolge_sayfalari[bolge_degerleri].append(ilkSatir)
            
        
        # İlk satırı kalın (bold) yap
        for cell in bolge_sayfalari[bolge_degerleri][1]:
            cell.font = Font(size=16, bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = openpyxl.styles.Border(left=openpyxl.styles.Side(style='thin'),
                                          right=openpyxl.styles.Side(style='thin'),
                                          top=openpyxl.styles.Side(style='thin'),
                                          bottom=openpyxl.styles.Side(style='thin'))
        # İlk satırın arka plan rengini gri yap
        for row in bolge_sayfalari[bolge_degerleri].iter_rows(min_row=1, max_row=1):
            for cell in row:
                cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

        # Her sütunun genişliğini arttır
        for column_index, column in enumerate(bolge_sayfalari[bolge_degerleri].columns, start=1):
            bolge_sayfalari[bolge_degerleri].column_dimensions[column[0].column_letter].width = kaynak_sheet.column_dimensions[kaynak_sheet.cell(row=1, column=column_index).column_letter].width + 2
    
    # Veriyi ilgili sayfaya ekle
    bolge_sayfalari[bolge_degerleri].append(list(satir))

silinecek_sheet_adi = "Sheet1"

if silinecek_sheet_adi in hedef_workbook.sheetnames:
    hedef_workbook.remove(hedef_workbook[silinecek_sheet_adi])
    print(f"{silinecek_sheet_adi} sayfası başarıyla silindi.")
else:
    print(f"{silinecek_sheet_adi} adında bir sayfa bulunamadı.")

hedef_workbook.save(hedef_dosya_adi)