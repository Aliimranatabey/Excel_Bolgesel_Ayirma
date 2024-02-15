import openpyxl

# Kaynak dosya adı ve hedef dosya adlarını tanımlayın
kaynak_dosya_adi = 'denemeMesh.xlsx'
hedef_dosya_adi = 'Mesh İhtiyacı Olabilecek Müşterilerin Dağılımları.xlsx'
# Kaynak dosyayı açın
kaynak_workbook = openpyxl.load_workbook(kaynak_dosya_adi)
hedef_workbook = openpyxl.load_workbook(hedef_dosya_adi)
hedef_workbook.active

# Kaynak dosyadaki çalışma sayfalarını alın
kaynak_sheets = kaynak_workbook.sheetnames

# Kaynak dosyadaki veri sayfasını seçin
kaynak_sheet = kaynak_workbook[kaynak_sheets[0]]

# Veri satırlarını okuyun
veri_satirlari = list(kaynak_sheet.iter_rows(values_only=True))

bolge_list=[]
a1,a2,a3,a4,a5,a6,a7,a8,a9,a10,a11=0,0,0,0,0,0,0,0,0,0,0
# Veri satırlarını dolaşarak her bölgeye uygun olanları ilgili listeye ekleyin
for satir in veri_satirlari[1:]:

    bolge_degerleri = satir[0]
    
    if bolge_degerleri not in bolge_list:
        bolge_list.append(bolge_degerleri)

    if bolge_degerleri==bolge_list[0]:
        if a1==0:
            bolge_degerleri_sheet1=hedef_workbook.create_sheet(title=bolge_degerleri)
            a1=a1+1
        bolge_degerleri_sheet1.append(list(satir))
    elif bolge_degerleri==bolge_list[1]:
        if a2==0:
            bolge_degerleri_sheet2=hedef_workbook.create_sheet(title=bolge_degerleri)
            a2=a2+1
        bolge_degerleri_sheet2.append(list(satir))
    elif bolge_degerleri==bolge_list[2]:
        if a3==0:
            bolge_degerleri_sheet3=hedef_workbook.create_sheet(title=bolge_degerleri)
            a3=a3+1
        bolge_degerleri_sheet3.append(list(satir))
    elif bolge_degerleri==bolge_list[3]:
        if a4==0:
            bolge_degerleri_sheet4=hedef_workbook.create_sheet(title=bolge_degerleri)
            a4=a4+1
        bolge_degerleri_sheet4.append(list(satir))
    elif bolge_degerleri==bolge_list[4]:
        if a5==0:
            bolge_degerleri_sheet5=hedef_workbook.create_sheet(title=bolge_degerleri)
            a5=a5+1
        bolge_degerleri_sheet5.append(list(satir))
    elif bolge_degerleri==bolge_list[5]:
        if a6==0:
            bolge_degerleri_sheet6=hedef_workbook.create_sheet(title=bolge_degerleri)
            a6=a6+1
        bolge_degerleri_sheet6.append(list(satir))
    elif bolge_degerleri==bolge_list[6]:
        if a7==0:
            bolge_degerleri_sheet7=hedef_workbook.create_sheet(title=bolge_degerleri)
            a7=a7+1
        bolge_degerleri_sheet7.append(list(satir))
    elif bolge_degerleri==bolge_list[7]:
        if a8==0:
            bolge_degerleri_sheet8=hedef_workbook.create_sheet(title=bolge_degerleri)
            a8=a8+1
        bolge_degerleri_sheet8.append(list(satir))
    elif bolge_degerleri==bolge_list[8]:
        if a9==0:
            bolge_degerleri_sheet9=hedef_workbook.create_sheet(title=bolge_degerleri)
            a9=a9+1
        bolge_degerleri_sheet9.append(list(satir))
    elif bolge_degerleri==bolge_list[9]:
        if a10==0:
            bolge_degerleri_sheet10=hedef_workbook.create_sheet(title=bolge_degerleri)
            a10=a10+1
        bolge_degerleri_sheet10.append(list(satir))
    elif bolge_degerleri==bolge_list[10]:
        if a11==0:
            bolge_degerleri_sheet11=hedef_workbook.create_sheet(title=bolge_degerleri)
            a11=a11+1
        bolge_degerleri_sheet11.append(list(satir))
hedef_workbook.save(hedef_dosya_adi)
        
