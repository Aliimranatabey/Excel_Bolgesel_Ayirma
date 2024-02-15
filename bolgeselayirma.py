import openpyxl

# Kaynak dosya adı ve hedef dosya adlarını tanımlayın
kaynak_dosya_adi = 'denemeMesh.xlsx'
hedef_dosya_adi = 'Mesh İhtiyacı Olabilecek Müşterilerin Dağılımları.xlsx'
# Kaynak dosyayı açın
kaynak_workbook = openpyxl.load_workbook(kaynak_dosya_adi)
hedef_workbook = openpyxl.load_workbook(hedef_dosya_adi)

# Kaynak dosyadaki çalışma sayfalarını alın
kaynak_sheets = kaynak_workbook.sheetnames

# Kaynak dosyadaki veri sayfasını seçin
kaynak_sheet = kaynak_workbook[kaynak_sheets[0]]

# Veri satırlarını okuyun
veri_satirlari = list(kaynak_sheet.iter_rows(values_only=True))
print(veri_satirlari)

# Boş bir dictionary oluşturun, her bölge için bir liste içerecek
#bolge_dict = {bolge: [] for bolge in veri_satirlari[0][0].split(",")}
bolge_dict={}
bolge_list=[]
#
b_deger=""
# Veri satırlarını dolaşarak her bölgeye uygun olanları ilgili listeye ekleyin
for satir in veri_satirlari[1:]:
    print(satir)
    bolge_degerleri = satir[0]
    print(bolge_degerleri)
    
    
    if bolge_degerleri not in bolge_list:
        bolge_list.append(bolge_degerleri)
        b_deger=bolge_degerleri
        action_list=[]
    
    action_list.append(satir)

    
    #bolge_dict[bolge_degerleri].append(action_list)
    
    
    #bolge_dict[bolge_degerleri].append(satir)

print(bolge_dict)
print(bolge_list)
print(action_list)

print(len(action_list))


"""
    for i, bölge in enumerate(bolge_dict.keys()):
        if bölge_değerleri[i] == bölge:  # Eğer bölge değeri 'x' ise
            bolge_dict[bölge].append(satir)
    

# Hedef dosyaları oluşturun ve ilgili bölgeye uygun satırları yazın
print(bolge_dict)
for bolge in bolge_dict.items():
    hedef_sheet = hedef_workbook[bolge]
    for satir in veri_satirlari:
        hedef_sheet.append(satir)
    hedef_workbook.save(f"{bolge}_Bolge.xlsx")

print("İşlem tamamlandı.")
"""


